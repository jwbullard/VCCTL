#!/usr/bin/env python3
"""
Advanced Export Service for VCCTL

Provides comprehensive export functionality including PDF reports, Excel spreadsheets,
JSON/XML data export, custom report templates, and batch operations.
"""

import logging
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Union, Callable
from dataclasses import asdict
import tempfile
import zipfile

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, KeepTogether
)
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie

# Excel generation
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.drawing.image import Image as ExcelImage

# XML processing
from lxml import etree

from .base_service import BaseService
from .file_operations_service import FileOperationsService
from ..models.cement import Cement
from ..models.aggregate import Aggregate
from ..visualization.plot_export import PlotExporter


class ExportFormat:
    """Supported export formats."""
    PDF = "pdf"
    EXCEL = "xlsx"
    JSON = "json"
    XML = "xml"
    CSV = "csv"
    ZIP = "zip"


class ReportTemplate:
    """Report template definition."""
    
    def __init__(self, name: str, description: str, sections: List[str]):
        self.name = name
        self.description = description
        self.sections = sections
        self.created_date = datetime.now()


class ExportService(BaseService):
    """Advanced export service for comprehensive data export and report generation."""
    
    def __init__(self, file_operations_service: FileOperationsService):
        super().__init__()
        self.file_operations_service = file_operations_service
        self.plot_exporter = PlotExporter()
        self.logger = logging.getLogger('VCCTL.Services.Export')
        
        # Report templates
        self.report_templates = {}
        self._initialize_templates()
        
        # Export statistics
        self.export_history = []
        
        self.logger.info("Export service initialized")
    
    def _initialize_templates(self):
        """Initialize default report templates."""
        
        # Comprehensive Project Report
        self.report_templates['comprehensive'] = ReportTemplate(
            name="Comprehensive Project Report",
            description="Complete project analysis with all parameters and results",
            sections=[
                "project_summary", "materials", "mix_design", "microstructure", 
                "hydration", "results", "plots", "conclusions"
            ]
        )
        
        # Materials Report
        self.report_templates['materials'] = ReportTemplate(
            name="Materials Analysis Report", 
            description="Focus on material properties and characteristics",
            sections=["project_summary", "materials", "plots"]
        )
        
        # Simulation Results Report
        self.report_templates['simulation'] = ReportTemplate(
            name="Simulation Results Report",
            description="Hydration and microstructure simulation results",
            sections=["project_summary", "hydration", "microstructure", "results", "plots"]
        )
        
        # Quick Summary Report
        self.report_templates['summary'] = ReportTemplate(
            name="Quick Summary Report",
            description="Brief overview of key parameters and results",
            sections=["project_summary", "results"]
        )
    
    # PDF Report Generation
    
    def generate_pdf_report(self, data: Dict[str, Any], output_path: Path,
                           template_name: str = 'comprehensive',
                           include_plots: bool = True) -> bool:
        """
        Generate comprehensive PDF report.
        
        Args:
            data: Project data to include in report
            output_path: Output file path
            template_name: Report template to use
            include_plots: Whether to include plot images
            
        Returns:
            Success status
        """
        try:
            if template_name not in self.report_templates:
                template_name = 'comprehensive'
            
            template = self.report_templates[template_name]
            
            # Create PDF document
            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build story
            story = []
            styles = getSampleStyleSheet()
            
            # Add title page
            story.extend(self._create_pdf_title_page(data, styles))
            
            # Add sections based on template
            for section in template.sections:
                if section == "project_summary":
                    story.extend(self._create_pdf_project_summary(data, styles))
                elif section == "materials":
                    story.extend(self._create_pdf_materials_section(data, styles))
                elif section == "mix_design":
                    story.extend(self._create_pdf_mix_design_section(data, styles))
                elif section == "microstructure":
                    story.extend(self._create_pdf_microstructure_section(data, styles))
                elif section == "hydration":
                    story.extend(self._create_pdf_hydration_section(data, styles))
                elif section == "results":
                    story.extend(self._create_pdf_results_section(data, styles))
                elif section == "plots" and include_plots:
                    story.extend(self._create_pdf_plots_section(data, styles))
                elif section == "conclusions":
                    story.extend(self._create_pdf_conclusions_section(data, styles))
            
            # Build PDF
            doc.build(story)
            
            self._log_export('PDF', output_path, template_name)
            self.logger.info(f"PDF report generated: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to generate PDF report: {e}")
            return False
    
    def _create_pdf_title_page(self, data: Dict[str, Any], styles) -> List:
        """Create PDF title page."""
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        project_name = data.get('project_name', 'VCCTL Project')
        story.append(Paragraph(f"VCCTL Analysis Report<br/>{project_name}", title_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Project information table
        project_info = [
            ['Project Name:', project_name],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['VCCTL Version:', data.get('vcctl_version', '1.0.0')],
            ['Author:', data.get('author', 'VCCTL User')]
        ]
        
        info_table = Table(project_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(info_table)
        story.append(PageBreak())
        
        return story
    
    def _create_pdf_project_summary(self, data: Dict[str, Any], styles) -> List:
        """Create project summary section."""
        story = []
        
        story.append(Paragraph("Project Summary", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        summary_text = data.get('project_description', 'No project description available.')
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 12))
        
        # Key parameters table
        if 'parameters' in data:
            params = data['parameters']
            param_data = [['Parameter', 'Value', 'Unit']]
            
            for key, value in params.items():
                if isinstance(value, dict) and 'value' in value and 'unit' in value:
                    param_data.append([
                        key.replace('_', ' ').title(),
                        str(value['value']),
                        value['unit']
                    ])
                else:
                    param_data.append([
                        key.replace('_', ' ').title(),
                        str(value),
                        ''
                    ])
            
            param_table = Table(param_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            param_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(Paragraph("Key Parameters", styles['Heading2']))
            story.append(Spacer(1, 6))
            story.append(param_table)
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_materials_section(self, data: Dict[str, Any], styles) -> List:
        """Create materials section."""
        story = []
        
        story.append(Paragraph("Materials", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        materials = data.get('materials', {})
        
        # Cement
        if 'cement' in materials:
            cement = materials['cement']
            story.append(Paragraph("Cement", styles['Heading2']))
            
            cement_data = [
                ['Property', 'Value'],
                ['Type', cement.get('cement_type', 'N/A')],
                ['Blaine Fineness', f"{cement.get('blaine_fineness', 'N/A')} m²/kg"],
                ['Density', f"{cement.get('density', 'N/A')} g/cm³"],
            ]
            
            # Add phase compositions
            phases = cement.get('phase_composition', {})
            for phase, percentage in phases.items():
                cement_data.append([f"{phase} Content", f"{percentage}%"])
            
            cement_table = Table(cement_data, colWidths=[2*inch, 2*inch])
            cement_table.setStyle(self._get_table_style())
            story.append(cement_table)
            story.append(Spacer(1, 12))
        
        # Aggregates
        if 'aggregates' in materials:
            story.append(Paragraph("Aggregates", styles['Heading2']))
            
            for i, agg in enumerate(materials['aggregates']):
                agg_data = [
                    ['Property', 'Value'],
                    ['Name', agg.get('name', f'Aggregate {i+1}')],
                    ['Type', agg.get('aggregate_type', 'N/A')],
                    ['Density', f"{agg.get('density', 'N/A')} g/cm³"],
                    ['Absorption', f"{agg.get('absorption', 'N/A')}%"],
                ]
                
                agg_table = Table(agg_data, colWidths=[2*inch, 2*inch])
                agg_table.setStyle(self._get_table_style())
                story.append(agg_table)
                story.append(Spacer(1, 6))
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_mix_design_section(self, data: Dict[str, Any], styles) -> List:
        """Create mix design section."""
        story = []
        
        story.append(Paragraph("Mix Design", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        mix_design = data.get('mix_design', {})
        
        if mix_design:
            mix_data = [['Component', 'Amount', 'Unit', 'Percentage']]
            
            total_mass = sum(float(v.get('amount', 0)) for v in mix_design.values() 
                           if isinstance(v, dict) and 'amount' in v)
            
            for component, details in mix_design.items():
                if isinstance(details, dict) and 'amount' in details:
                    amount = float(details['amount'])
                    unit = details.get('unit', 'kg/m³')
                    percentage = (amount / total_mass * 100) if total_mass > 0 else 0
                    
                    mix_data.append([
                        component.replace('_', ' ').title(),
                        f"{amount:.2f}",
                        unit,
                        f"{percentage:.1f}%"
                    ])
            
            mix_table = Table(mix_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch])
            mix_table.setStyle(self._get_table_style())
            story.append(mix_table)
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_microstructure_section(self, data: Dict[str, Any], styles) -> List:
        """Create microstructure section."""
        story = []
        
        story.append(Paragraph("Microstructure Parameters", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        microstructure = data.get('microstructure', {})
        
        if microstructure:
            micro_data = [['Parameter', 'Value', 'Unit']]
            
            for param, value in microstructure.items():
                if isinstance(value, dict):
                    micro_data.append([
                        param.replace('_', ' ').title(),
                        str(value.get('value', 'N/A')),
                        value.get('unit', '')
                    ])
                else:
                    micro_data.append([
                        param.replace('_', ' ').title(),
                        str(value),
                        ''
                    ])
            
            micro_table = Table(micro_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            micro_table.setStyle(self._get_table_style())
            story.append(micro_table)
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_hydration_section(self, data: Dict[str, Any], styles) -> List:
        """Create hydration section."""
        story = []
        
        story.append(Paragraph("Hydration Parameters", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        hydration = data.get('hydration', {})
        
        if hydration:
            hydration_data = [['Parameter', 'Value', 'Unit']]
            
            for param, value in hydration.items():
                if isinstance(value, dict):
                    hydration_data.append([
                        param.replace('_', ' ').title(),
                        str(value.get('value', 'N/A')),
                        value.get('unit', '')
                    ])
                else:
                    hydration_data.append([
                        param.replace('_', ' ').title(),
                        str(value),
                        ''
                    ])
            
            hydration_table = Table(hydration_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            hydration_table.setStyle(self._get_table_style())
            story.append(hydration_table)
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_results_section(self, data: Dict[str, Any], styles) -> List:
        """Create results section."""
        story = []
        
        story.append(Paragraph("Results", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        results = data.get('results', {})
        
        if results:
            results_data = [['Result', 'Value', 'Unit']]
            
            for result, value in results.items():
                if isinstance(value, dict):
                    results_data.append([
                        result.replace('_', ' ').title(),
                        str(value.get('value', 'N/A')),
                        value.get('unit', '')
                    ])
                else:
                    results_data.append([
                        result.replace('_', ' ').title(),
                        str(value),
                        ''
                    ])
            
            results_table = Table(results_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
            results_table.setStyle(self._get_table_style())
            story.append(results_table)
        
        story.append(PageBreak())
        return story
    
    def _create_pdf_plots_section(self, data: Dict[str, Any], styles) -> List:
        """Create plots section."""
        story = []
        
        story.append(Paragraph("Plots and Visualizations", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        # This would include plot images if available
        plot_paths = data.get('plot_paths', [])
        
        for plot_path in plot_paths:
            if Path(plot_path).exists():
                try:
                    # Add plot image
                    img = Image(plot_path, width=6*inch, height=4*inch)
                    story.append(img)
                    story.append(Spacer(1, 12))
                except Exception as e:
                    self.logger.warning(f"Could not include plot {plot_path}: {e}")
        
        if not plot_paths:
            story.append(Paragraph("No plots available for inclusion.", styles['Normal']))
        
        return story
    
    def _create_pdf_conclusions_section(self, data: Dict[str, Any], styles) -> List:
        """Create conclusions section."""
        story = []
        
        story.append(Paragraph("Conclusions", styles['Heading1']))
        story.append(Spacer(1, 12))
        
        conclusions = data.get('conclusions', 'No conclusions provided.')
        story.append(Paragraph(conclusions, styles['Normal']))
        
        return story
    
    def _get_table_style(self) -> TableStyle:
        """Get standard table style."""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
    
    # Excel Export Functions
    
    def generate_excel_report(self, data: Dict[str, Any], output_path: Path,
                             include_charts: bool = True) -> bool:
        """
        Generate comprehensive Excel report with multiple sheets.
        
        Args:
            data: Project data to export
            output_path: Output file path
            include_charts: Whether to include charts
            
        Returns:
            Success status
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_excel_summary_sheet(wb, data)
            self._create_excel_materials_sheet(wb, data)
            self._create_excel_mix_design_sheet(wb, data)
            self._create_excel_parameters_sheet(wb, data)
            self._create_excel_results_sheet(wb, data)
            
            if include_charts:
                self._create_excel_charts_sheet(wb, data)
            
            # Save workbook
            wb.save(str(output_path))
            
            self._log_export('Excel', output_path, 'comprehensive')
            self.logger.info(f"Excel report generated: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            return False
    
    def _create_excel_summary_sheet(self, wb, data: Dict[str, Any]):
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet("Summary")
        
        # Headers
        header_font = Font(bold=True, size=14)
        fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        ws['A1'] = "VCCTL Project Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Project information
        row = 3
        info_items = [
            ('Project Name:', data.get('project_name', 'N/A')),
            ('Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('VCCTL Version:', data.get('vcctl_version', '1.0.0')),
            ('Author:', data.get('author', 'VCCTL User'))
        ]
        
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_excel_materials_sheet(self, wb, data: Dict[str, Any]):
        """Create materials sheet in Excel workbook."""
        ws = wb.create_sheet("Materials")
        
        materials = data.get('materials', {})
        
        row = 1
        header_font = Font(bold=True)
        
        # Cement section
        if 'cement' in materials:
            cement = materials['cement']
            
            ws[f'A{row}'] = "Cement Properties"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            ws[f'A{row}'] = "Property"
            ws[f'B{row}'] = "Value"
            for cell in [ws[f'A{row}'], ws[f'B{row}']]:
                cell.font = header_font
            row += 1
            
            for prop, value in cement.items():
                ws[f'A{row}'] = prop.replace('_', ' ').title()
                ws[f'B{row}'] = str(value)
                row += 1
            
            row += 2
        
        # Aggregates section
        if 'aggregates' in materials:
            ws[f'A{row}'] = "Aggregates"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            for i, agg in enumerate(materials['aggregates']):
                ws[f'A{row}'] = f"Aggregate {i+1}"
                ws[f'A{row}'].font = header_font
                row += 1
                
                for prop, value in agg.items():
                    ws[f'A{row}'] = prop.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
                
                row += 1
    
    def _create_excel_mix_design_sheet(self, wb, data: Dict[str, Any]):
        """Create mix design sheet in Excel workbook."""
        ws = wb.create_sheet("Mix Design")
        
        mix_design = data.get('mix_design', {})
        
        if mix_design:
            # Headers
            ws['A1'] = "Component"
            ws['B1'] = "Amount"
            ws['C1'] = "Unit"
            ws['D1'] = "Percentage"
            
            for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
                cell.font = Font(bold=True)
            
            # Calculate total for percentages
            total_mass = sum(float(v.get('amount', 0)) for v in mix_design.values() 
                           if isinstance(v, dict) and 'amount' in v)
            
            row = 2
            for component, details in mix_design.items():
                if isinstance(details, dict) and 'amount' in details:
                    amount = float(details['amount'])
                    unit = details.get('unit', 'kg/m³')
                    percentage = (amount / total_mass * 100) if total_mass > 0 else 0
                    
                    ws[f'A{row}'] = component.replace('_', ' ').title()
                    ws[f'B{row}'] = amount
                    ws[f'C{row}'] = unit
                    ws[f'D{row}'] = f"{percentage:.1f}%"
                    row += 1
    
    def _create_excel_parameters_sheet(self, wb, data: Dict[str, Any]):
        """Create parameters sheet in Excel workbook."""
        ws = wb.create_sheet("Parameters")
        
        # Headers
        ws['A1'] = "Parameter"
        ws['B1'] = "Value"
        ws['C1'] = "Unit"
        
        for cell in [ws['A1'], ws['B1'], ws['C1']]:
            cell.font = Font(bold=True)
        
        row = 2
        
        # Microstructure parameters
        microstructure = data.get('microstructure', {})
        if microstructure:
            ws[f'A{row}'] = "Microstructure Parameters"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for param, value in microstructure.items():
                ws[f'A{row}'] = param.replace('_', ' ').title()
                if isinstance(value, dict):
                    ws[f'B{row}'] = value.get('value', 'N/A')
                    ws[f'C{row}'] = value.get('unit', '')
                else:
                    ws[f'B{row}'] = str(value)
                row += 1
        
        # Hydration parameters
        hydration = data.get('hydration', {})
        if hydration:
            row += 1
            ws[f'A{row}'] = "Hydration Parameters"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for param, value in hydration.items():
                ws[f'A{row}'] = param.replace('_', ' ').title()
                if isinstance(value, dict):
                    ws[f'B{row}'] = value.get('value', 'N/A')
                    ws[f'C{row}'] = value.get('unit', '')
                else:
                    ws[f'B{row}'] = str(value)
                row += 1
    
    def _create_excel_results_sheet(self, wb, data: Dict[str, Any]):
        """Create results sheet in Excel workbook."""
        ws = wb.create_sheet("Results")
        
        results = data.get('results', {})
        
        if results:
            # Headers
            ws['A1'] = "Result"
            ws['B1'] = "Value"
            ws['C1'] = "Unit"
            
            for cell in [ws['A1'], ws['B1'], ws['C1']]:
                cell.font = Font(bold=True)
            
            row = 2
            for result, value in results.items():
                ws[f'A{row}'] = result.replace('_', ' ').title()
                if isinstance(value, dict):
                    ws[f'B{row}'] = value.get('value', 'N/A')
                    ws[f'C{row}'] = value.get('unit', '')
                else:
                    ws[f'B{row}'] = str(value)
                row += 1
    
    def _create_excel_charts_sheet(self, wb, data: Dict[str, Any]):
        """Create charts sheet in Excel workbook."""
        ws = wb.create_sheet("Charts")
        
        # Placeholder for chart creation
        ws['A1'] = "Charts and Visualizations"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Note: Actual chart creation would require specific data series
        ws['A3'] = "Chart data would be generated here based on simulation results"
    
    # JSON/XML Export Functions
    
    def export_json(self, data: Dict[str, Any], output_path: Path,
                   indent: int = 2) -> bool:
        """
        Export data to JSON format.
        
        Args:
            data: Data to export
            output_path: Output file path
            indent: JSON indentation
            
        Returns:
            Success status
        """
        try:
            # Convert any non-serializable objects
            serializable_data = self._make_json_serializable(data)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(serializable_data, f, indent=indent, ensure_ascii=False)
            
            self._log_export('JSON', output_path)
            self.logger.info(f"JSON export completed: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export JSON: {e}")
            return False
    
    def export_xml(self, data: Dict[str, Any], output_path: Path,
                  root_tag: str = 'vcctl_project') -> bool:
        """
        Export data to XML format.
        
        Args:
            data: Data to export
            output_path: Output file path
            root_tag: Root XML element name
            
        Returns:
            Success status
        """
        try:
            root = ET.Element(root_tag)
            
            # Add metadata
            metadata = ET.SubElement(root, 'metadata')
            ET.SubElement(metadata, 'generated').text = datetime.now().isoformat()
            ET.SubElement(metadata, 'vcctl_version').text = data.get('vcctl_version', '1.0.0')
            
            # Convert data to XML
            self._dict_to_xml(data, root)
            
            # Create tree and write
            tree = ET.ElementTree(root)
            ET.indent(tree, space="  ", level=0)
            tree.write(output_path, encoding='utf-8', xml_declaration=True)
            
            self._log_export('XML', output_path)
            self.logger.info(f"XML export completed: {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to export XML: {e}")
            return False
    
    def _dict_to_xml(self, data: Dict[str, Any], parent: ET.Element):
        """Convert dictionary to XML elements recursively."""
        for key, value in data.items():
            # Clean key for XML
            clean_key = str(key).replace(' ', '_').replace('-', '_')
            
            if isinstance(value, dict):
                element = ET.SubElement(parent, clean_key)
                self._dict_to_xml(value, element)
            elif isinstance(value, list):
                for item in value:
                    element = ET.SubElement(parent, clean_key)
                    if isinstance(item, dict):
                        self._dict_to_xml(item, element)
                    else:
                        element.text = str(item)
            else:
                element = ET.SubElement(parent, clean_key)
                element.text = str(value)
    
    # Batch Export Functions
    
    def batch_export(self, data: Dict[str, Any], output_dir: Path,
                    formats: List[str] = None, template_name: str = 'comprehensive') -> Dict[str, bool]:
        """
        Perform batch export to multiple formats.
        
        Args:
            data: Data to export
            output_dir: Output directory
            formats: List of formats to export to
            template_name: Report template to use
            
        Returns:
            Dictionary of format -> success status
        """
        if formats is None:
            formats = [ExportFormat.PDF, ExportFormat.EXCEL, ExportFormat.JSON]
        
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        results = {}
        base_name = data.get('project_name', 'vcctl_export').replace(' ', '_')
        
        for format_type in formats:
            try:
                if format_type == ExportFormat.PDF:
                    output_path = output_dir / f"{base_name}.pdf"
                    results[format_type] = self.generate_pdf_report(
                        data, output_path, template_name
                    )
                
                elif format_type == ExportFormat.EXCEL:
                    output_path = output_dir / f"{base_name}.xlsx"
                    results[format_type] = self.generate_excel_report(data, output_path)
                
                elif format_type == ExportFormat.JSON:
                    output_path = output_dir / f"{base_name}.json"
                    results[format_type] = self.export_json(data, output_path)
                
                elif format_type == ExportFormat.XML:
                    output_path = output_dir / f"{base_name}.xml"
                    results[format_type] = self.export_xml(data, output_path)
                
                elif format_type == ExportFormat.ZIP:
                    results[format_type] = self._create_export_archive(
                        data, output_dir, base_name
                    )
                
                else:
                    self.logger.warning(f"Unsupported export format: {format_type}")
                    results[format_type] = False
                    
            except Exception as e:
                self.logger.error(f"Failed to export {format_type}: {e}")
                results[format_type] = False
        
        return results
    
    def _create_export_archive(self, data: Dict[str, Any], output_dir: Path,
                              base_name: str) -> bool:
        """Create ZIP archive with all export formats."""
        try:
            zip_path = output_dir / f"{base_name}_complete.zip"
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Create temporary files for all formats
                with tempfile.TemporaryDirectory() as temp_dir:
                    temp_path = Path(temp_dir)
                    
                    # Generate all formats
                    pdf_path = temp_path / f"{base_name}.pdf"
                    excel_path = temp_path / f"{base_name}.xlsx"
                    json_path = temp_path / f"{base_name}.json"
                    xml_path = temp_path / f"{base_name}.xml"
                    
                    # Generate files
                    self.generate_pdf_report(data, pdf_path)
                    self.generate_excel_report(data, excel_path)
                    self.export_json(data, json_path)
                    self.export_xml(data, xml_path)
                    
                    # Add to archive
                    for file_path in [pdf_path, excel_path, json_path, xml_path]:
                        if file_path.exists():
                            zf.write(file_path, file_path.name)
            
            self._log_export('ZIP', zip_path)
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to create export archive: {e}")
            return False
    
    # Utility Functions
    
    def _make_json_serializable(self, obj: Any) -> Any:
        """Make object JSON serializable."""
        if isinstance(obj, dict):
            return {key: self._make_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif hasattr(obj, '__dict__'):
            return self._make_json_serializable(asdict(obj) if hasattr(obj, '__dataclass_fields__') else obj.__dict__)
        elif isinstance(obj, (datetime,)):
            return obj.isoformat()
        elif isinstance(obj, Path):
            return str(obj)
        else:
            try:
                json.dumps(obj)
                return obj
            except (TypeError, ValueError):
                return str(obj)
    
    def _log_export(self, format_type: str, output_path: Path, template: str = None):
        """Log export operation."""
        export_record = {
            'timestamp': datetime.now().isoformat(),
            'format': format_type,
            'output_path': str(output_path),
            'template': template,
            'size_bytes': output_path.stat().st_size if output_path.exists() else 0
        }
        
        self.export_history.append(export_record)
        
        # Keep only last 100 exports in memory
        if len(self.export_history) > 100:
            self.export_history = self.export_history[-100:]
    
    def get_export_history(self) -> List[Dict[str, Any]]:
        """Get export history."""
        return self.export_history.copy()
    
    def get_available_templates(self) -> Dict[str, ReportTemplate]:
        """Get available report templates."""
        return self.report_templates.copy()
    
    def add_custom_template(self, template: ReportTemplate) -> bool:
        """Add custom report template."""
        try:
            self.report_templates[template.name.lower().replace(' ', '_')] = template
            self.logger.info(f"Added custom template: {template.name}")
            return True
        except Exception as e:
            self.logger.error(f"Failed to add custom template: {e}")
            return False